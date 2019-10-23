from predict_each_m_v.classify import preprocess_handle
from openpyxl import load_workbook
from keras.models import Sequential
from sklearn import preprocessing
import pandas as pd
from keras.layers import LSTM, Dense, Activation
import matplotlib.pyplot as plt
import numpy as np
from keras.layers import *
from keras.callbacks import Callback, ModelCheckpoint
from sklearn.svm import SVR
from sklearn.metrics import mean_squared_error

def svr_base_model(train_x, test_x, train_y, test_y, org_x, org_y):
    regressor = SVR(kernel='rbf', C=100, gamma=0.01)
    regressor.fit(train_x, train_y)
    pred_y = regressor.predict(test_x)
    mse = mean_squared_error(test_y, pred_y)
    print('MSE', mse)
    pre_y = regressor.predict(org_x)
    # 这里算整体的MSE会很大，因为没有归一化，我的y值太大了
    def mape(y_true, y_pred):
        return np.mean(np.abs((y_pred - y_true) / y_true)) * 100
    # mse2 = mean_squared_error(org_y, pre_y)
    # print('Total Mse', mse2)
    # print(pre_y)
    max_ = org_y.argmax()
    # print(y[max_])
    min_ = org_y.argmin()
    # print(y[min_])
    res_list = []
    for list_ in pre_y:
        transfer = (org_y[max_] - org_y[min_]) * list_ + org_y[min_]
        # print(transfer)
        res_list.append(round(transfer, 4))
    # print(res_list)
    plt.plot(np.array(org_y), 'blue', label='true data')
    plt.plot(np.array(res_list), 'r', label='train data')
    mape_s = mape(org_y, np.array(res_list))
    print('MAPE', mape_s)
    plt.show()


def scaler_adjust(data, scaler_model):
    if scaler_model == 1:
        # 归一化到[0,1]区间
        # x = data[]
        min_max_scaler = preprocessing.MinMaxScaler(feature_range=(0, 1))
        data_nor = min_max_scaler.fit_transform(data)
        # print(x_nor[-1])
        return data_nor, min_max_scaler
    # elif scaler_model == 2:
    #     # 归一化到[-1, 1]区间
    #     df = pd.DataFrame(x_data)
    #     # print(df.describe())
    #     # print(df)
    #     # df_norm = (df - df.min()) / (df.max() - df.min())
    #     # print(df_norm)
    #     df_norm = (2*(df - df.min())/(df.max() - df.min()))-1
    #     # print(df_norm)
    #     return np.array(df_norm)
    # elif scaler_model == 3:
    #     # 归一化到[-1, 0]区间
    #     df = pd.DataFrame(x_data)
    #     # print(df.describe())
    #     # print(df)
    #     # df_norm = (df - df.min()) / (df.max() - df.min())
    #     # print(df_norm)
    #     df_norm = (1*(df - df.min())/(df.max() - df.min()))-1
    #     # print(df_norm)
    #     return np.array(df_norm)
    # elif scaler_model == 4:
    #     # Z-score成标准正态
    #     df = pd.DataFrame(x_data)
    #     df_norm = (df - df.mean()) / (df.std())
    #     # print(df_norm)
    #     return np.array(df_norm)
    # else:
    #     return np.array(x_data)

def load_data(scaler_model):
    df = pd.read_excel('org_regression_data.xlsx', sheet_name='Sheet1')  # 可以通过sheet_name来指定读取的表单
    data_nor, scaler = scaler_adjust(df.iloc[:, 1:20], scaler_model)
    # df = pd.DataFrame(data_nor).corr()
    # import seaborn as sns
    # sns.heatmap(df, vmax=0.8, cmap='RdBu_r', square=True, center=0)
    # plt.savefig('./BluesStateRelation.png')
    # plt.show()
    # 经过标准化处理后数据变为np.array形式
    x_nor = data_nor[:, 0:18]
    # debug here
    # print(x_nor[-1, 0:18])
    y_nor = data_nor[:, -1]
    return x_nor, y_nor, df['沪深300'], scaler


def build_model():
    # input_dim是输入的train_x的最后一个维度，train_x的维度为(n_samples, time_steps, input_dim)
    max_sentence_length = 16
    max_features = 100
    batch_size = 8
    embedding_size = 64
    hidden_layer_size = 64
    period = 100
    classes = 2
    dropuot = .3
    epoch = 10000
    model = Sequential()
    # 先简单的直接用单个值进行预测
    model.add(LSTM(units=256, input_shape=(None, 18), return_sequences=True))
    model.add(LSTM(units=256))
    # model.add(Embedding(max_features, embedding_size, input_length=max_sentence_length))
    model.add(Dropout(0.5))
    # model.add(Bidirectional(CuDNNLSTM(128)))
    model.add(Dense(1, activation='relu'))

    # model.add(Activation('linear'))
    # model.add(Activation('sigmoid'))

    model.compile(loss='mse', optimizer='Adam')
    model.summary()
    return model



def train_model(train_x, test_x, train_y, test_y):
    # 终于找到key problem了，LSTM的输入应该为[batchsize， 【用多少天的时间进行下一天的预测】， 变量个数]
    batch_size = 8
    epoch = 1000
    model = build_model()

    checkpoint = ModelCheckpoint(filepath='./models/{epoch:02d}-{val_loss:.2f}_Reg5_LSTM_model_20v.h5',
                                 monitor='val_loss', save_best_only=True, save_weights_only=True, mode='min',
                                 period=1)
    # print(train_x)
    # print(train_y)
    re_train = train_x.reshape((train_x.shape[0], 1, train_x.shape[1]))
    re_test = test_x.reshape((test_x.shape[0], 1, test_x.shape[1]))
    history = model.fit(re_train, train_y, batch_size=batch_size, epochs=epoch, validation_data=(re_test, test_y),
                callbacks=[checkpoint])
    plt.plot(history.history['loss'], label='train')
    plt.plot(history.history['val_loss'], label='test')
    plt.legend()
    plt.show()

def plot_test(x, org_y):
    model = Sequential()
    # 先简单的直接用单个值进行预测
    model.add(LSTM(units=256, input_shape=(None, 18), return_sequences=True))
    model.add(LSTM(units=256))
    model.add(Dropout(0.5))
    model.add(Dense(1, activation='relu'))
    # 579-0.00_Reg_LSTM_model 12-0.01_Reg_LSTM_model 528-0.00_Reg5_LSTM_model_20v  306-0.00_Reg5_LSTM_model_20v
    model.load_weights('./models/528-0.00_Reg5_LSTM_model_20v.h5')
    model.compile(loss='mse', optimizer='adam')
    plt.plot(np.array(org_y), 'blue', label='true data')
    re_x = x.reshape((x.shape[0], 1, x.shape[1]))
    com_list = model.predict(re_x).tolist()
    res_list = reshape_org_shape(org_y, com_list)
    plt.plot(res_list, 'r', label='predict')
    plt.show()

def reshape_org_shape(y, com_list):
    max_ = y.argmax()
    # print(y[max_])
    min_ = y.argmin()
    # print(y[min_])
    res_list = []
    for list_ in com_list:
        transfer = (y[max_] - y[min_]) * list_[0] + y[min_]
        # print(transfer)
        res_list.append([round(transfer, 4)])
    # print(res_list)
    return res_list


def statistic():
    wb = load_workbook('org_regression_data.xlsx')
    ws = wb.get_sheet_by_name('statistic')
    v = 0
    z = 0
    zt = 0
    f = 0
    ft = 0
    for i in range(ws.max_row+1):
        gt = ws.cell(row=i+2, column=2).value
        pred = ws.cell(row=i+2, column=3).value
        if gt == 1:
           zt += 1
        elif gt == -1:
            ft += 1
        else:
            continue
        if gt == pred:
            v += 1
            if gt == 1:
                z += 1
            else:
                f += 1
        else:
            continue
    print('True num', v)
    print('True ratio', v/(ws.max_row-1))
    print('Pos num', z)
    print('Pos ratio', z/zt)
    print('Neg num', f)
    print('Neg ratio', f/ft)

def pre_data(x, y, cur_x_data):
    wb = load_workbook('org_regression_data.xlsx')
    ws = wb.get_sheet_by_name('each_predict')
    model = Sequential()
    # 先简单的直接用单个值进行预测
    model.add(LSTM(units=256, input_shape=(None, 18), return_sequences=True))
    model.add(LSTM(units=256))
    model.add(Dropout(0.5))
    model.add(Dense(1, activation='relu'))
    # 12-0.01_Reg_LSTM_model 579-0.00_Reg_LSTM_model
    model.load_weights('./models/528-0.00_Reg5_LSTM_model_20v.h5')
    model.compile(loss='mse', optimizer='adam')
    plt.plot(np.array(y), 'blue', label='true data')
    for index, v in enumerate(y):
        ws.cell(row=index+2, column=2, value=v)
    re_x = x.reshape((x.shape[0], 1, x.shape[1]))
    # print(re_x)
    cur_pre = model.predict(np.array(cur_x_data)).tolist()
    # print(cur_pre[0])
    com_list = model.predict(re_x).tolist()
    # print(len(com_list))
    com_list.append(cur_pre[0])
    # print(len(com_list))
    res_array = np.array(com_list)
    res_list = reshape_org_shape(y, com_list)
    print('Cur predict value:', res_list[-1][0])
    plt.plot(np.array(res_list), 'r', label='train data')
    # plt.plot(len(res_list), res_list[-1][0], 'y', label='cur predict')
    plt.show()
    for index, v in enumerate(res_list):
        ws.cell(row=index + 2, column=3, value=v[0])
    wb.save('org_regression_data.xlsx')

if __name__ == '__main__':
    scaler_model = 1
    x, y, org_y, scaler = load_data(scaler_model)
    train_x, test_x, train_y, test_y = preprocess_handle(x, y)
    svr_base_model(train_x, test_x, train_y, test_y, x, org_y)
    # train_model(train_x, test_x, train_y, test_y)
    # plot_test(x, org_y)
    # 目前想预测下一个的值，只能把所有数据先利用debug进行归一化，接下来可以开发一个自动进行归一化的函数,当前为9月份值，进行预测10月份沪深300
    cur_x_data = [[[0.17213226, 0.99624902, 0.47892745, 0.13020543, 0.98528667, 0.08946666,
                     0.21514927, 0.21527661, 0.45385555, 0.48040999, 0.47582535, 0.48798801,
                     0.09583032, 0.04683018, 0.53921569, 0.2281357,  1.,         0.40525739]]]
    # 对最新的结果进行预测，输出预测的下月结果，并在excel中保存数据
    # pre_data(x, org_y, cur_x_data)

    # 查看当前模型预测真实拐点的概率
    statistic()